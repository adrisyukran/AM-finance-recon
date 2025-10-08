"""
Exporter Module
Handles file export operations for reconciliation results
"""

import os
import pandas as pd
from typing import Dict, List, Tuple, Optional
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class Exporter:
    """Export reconciliation results to various formats"""

    def __init__(self, config):
        """
        Initialize Exporter with configuration

        Args:
            config: Application configuration object
        """
        self.config = config
        self.upload_folder = config.UPLOAD_FOLDER

    def create_new_reconciled_file(
        self,
        df: pd.DataFrame,
        matches: List[Dict],
        output_path: str,
        file_format: str = "xlsx",
    ) -> Tuple[bool, str]:
        """
        Create a new file with grouped matched transactions

        Args:
            df: Original DataFrame with all transactions
            matches: List of matched groups
            output_path: Path where file should be saved
            file_format: Output format ('csv' or 'xlsx')

        Returns:
            Tuple of (success, message)
        """
        try:
            # Create reconciliation DataFrame
            reconciled_data = self._prepare_grouped_data(df, matches)

            if file_format.lower() == "csv":
                return self._export_grouped_csv(reconciled_data, output_path)
            elif file_format.lower() in ["xlsx", "xls"]:
                return self._export_grouped_excel(reconciled_data, matches, output_path)
            else:
                return False, f"Unsupported format: {file_format}"

        except Exception as e:
            return False, f"Error creating reconciled file: {str(e)}"

    def update_existing_file(
        self,
        original_file_path: str,
        df: pd.DataFrame,
        matches: List[Dict],
        options: Dict,
    ) -> Tuple[bool, str, str]:
        """
        Update existing file with reconciliation status

        Args:
            original_file_path: Path to original file
            df: DataFrame with reconciliation data
            matches: List of matched groups
            options: Export options dictionary containing:
                - status_columns: List of column names to add/update
                - status_text: Text to add to matched rows
                - highlight: Boolean to highlight matched rows

        Returns:
            Tuple of (success, message, output_path)
        """
        try:
            # Generate output filename
            base_name, ext = os.path.splitext(original_file_path)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"{base_name}_reconciled_{timestamp}{ext}"

            # Update DataFrame with status columns
            updated_df = self._add_status_columns(df, matches, options)

            if ext.lower() == ".csv":
                updated_df.to_csv(output_path, index=False)
                return True, f"File updated successfully", output_path

            elif ext.lower() in [".xlsx", ".xls"]:
                # Export to Excel with optional highlighting
                updated_df.to_excel(output_path, index=False, engine="openpyxl")

                if options.get("highlight", False):
                    self._apply_highlighting(output_path, updated_df, matches)

                return True, f"File updated successfully", output_path

            else:
                return False, "Unsupported file format", ""

        except Exception as e:
            return False, f"Error updating file: {str(e)}", ""

    def _prepare_grouped_data(
        self, df: pd.DataFrame, matches: List[Dict]
    ) -> pd.DataFrame:
        """
        Prepare data grouped by match groups

        Args:
            df: Original DataFrame
            matches: List of matched groups

        Returns:
            DataFrame with grouped reconciliation data
        """
        grouped_rows = []

        for idx, match in enumerate(matches):
            match_group_id = f"MG_{idx:04d}"
            revenue = match.get("revenue", {})
            expenses = match.get("expenses", [])
            confidence = match.get("confidence", 0.0)
            match_type = match.get("match_type", "unknown")

            # Add revenue row
            if revenue:
                revenue_row = revenue.copy()
                revenue_row["match_group_id"] = match_group_id
                revenue_row["match_status"] = "matched"
                revenue_row["match_confidence"] = f"{confidence * 100:.1f}%"
                revenue_row["match_type"] = match_type
                revenue_row["group_position"] = "revenue"
                grouped_rows.append(revenue_row)

            # Add expense rows
            for exp_num, expense in enumerate(expenses, 1):
                expense_row = expense.copy()
                expense_row["match_group_id"] = match_group_id
                expense_row["match_status"] = "matched"
                expense_row["match_confidence"] = f"{confidence * 100:.1f}%"
                expense_row["match_type"] = match_type
                expense_row["group_position"] = f"expense_{exp_num}"
                grouped_rows.append(expense_row)

            # Add separator row for better readability
            if idx < len(matches) - 1:
                separator = {col: "" for col in grouped_rows[0].keys()}
                separator["match_group_id"] = "---"
                grouped_rows.append(separator)

        # Add unmatched transactions
        unmatched = df[df["match_status"] == "unmatched"]
        for _, row in unmatched.iterrows():
            unmatched_row = row.to_dict()
            unmatched_row["match_group_id"] = "UNMATCHED"
            unmatched_row["match_status"] = "unmatched"
            unmatched_row["match_confidence"] = "N/A"
            unmatched_row["match_type"] = "N/A"
            unmatched_row["group_position"] = "unmatched"
            grouped_rows.append(unmatched_row)

        return pd.DataFrame(grouped_rows)

    def _export_grouped_csv(
        self, df: pd.DataFrame, output_path: str
    ) -> Tuple[bool, str]:
        """
        Export grouped data to CSV

        Args:
            df: DataFrame to export
            output_path: Output file path

        Returns:
            Tuple of (success, message)
        """
        try:
            df.to_csv(output_path, index=False)
            return True, f"CSV file created successfully: {output_path}"
        except Exception as e:
            return False, f"Error exporting to CSV: {str(e)}"

    def _export_grouped_excel(
        self, df: pd.DataFrame, matches: List[Dict], output_path: str
    ) -> Tuple[bool, str]:
        """
        Export grouped data to Excel with formatting

        Args:
            df: DataFrame to export
            matches: List of matched groups
            output_path: Output file path

        Returns:
            Tuple of (success, message)
        """
        try:
            # Write to Excel
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Reconciliation", index=False)

                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets["Reconciliation"]

                # Apply formatting
                self._format_grouped_excel(worksheet, df, matches)

            return True, f"Excel file created successfully: {output_path}"

        except Exception as e:
            return False, f"Error exporting to Excel: {str(e)}"

    def _format_grouped_excel(self, worksheet, df: pd.DataFrame, matches: List[Dict]):
        """
        Apply formatting to grouped Excel worksheet

        Args:
            worksheet: openpyxl worksheet object
            df: DataFrame being exported
            matches: List of matched groups
        """
        # Define styles
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        matched_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        unmatched_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )
        separator_fill = PatternFill(
            start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Format header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Format data rows
        for idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            match_status = df.iloc[idx - 2].get("match_status", "")

            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

                # Apply background color based on status
                if match_status == "matched":
                    cell.fill = matched_fill
                elif match_status == "unmatched":
                    cell.fill = unmatched_fill
                elif df.iloc[idx - 2].get("match_group_id", "") == "---":
                    cell.fill = separator_fill

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _add_status_columns(
        self, df: pd.DataFrame, matches: List[Dict], options: Dict
    ) -> pd.DataFrame:
        """
        Add status columns to DataFrame

        Args:
            df: Original DataFrame
            matches: List of matched groups
            options: Export options

        Returns:
            Updated DataFrame
        """
        df_copy = df.copy()

        status_columns = options.get("status_columns", [])
        status_text = options.get("status_text", self.config.DEFAULT_STATUS_TEXT)

        # Add new columns if they don't exist
        for col in status_columns:
            if col not in df_copy.columns:
                df_copy[col] = ""

        # Update matched rows
        matched_indices = df_copy[df_copy["match_status"] == "matched"].index

        for idx in matched_indices:
            for col in status_columns:
                df_copy.at[idx, col] = status_text

        return df_copy

    def _apply_highlighting(
        self, file_path: str, df: pd.DataFrame, matches: List[Dict]
    ):
        """
        Apply highlighting to matched rows in Excel file

        Args:
            file_path: Path to Excel file
            df: DataFrame with data
            matches: List of matched groups
        """
        try:
            # Load workbook
            workbook = load_workbook(file_path)
            worksheet = workbook.active

            # Define highlight color
            yellow_fill = PatternFill(
                start_color=self.config.DEFAULT_HIGHLIGHT_COLOR,
                end_color=self.config.DEFAULT_HIGHLIGHT_COLOR,
                fill_type="solid",
            )

            # Highlight matched rows
            matched_indices = df[df["match_status"] == "matched"].index

            for idx in matched_indices:
                excel_row = idx + 2  # +2 for Excel 1-indexing and header row

                for cell in worksheet[excel_row]:
                    cell.fill = yellow_fill

            # Save workbook
            workbook.save(file_path)

        except Exception as e:
            print(f"Warning: Could not apply highlighting: {str(e)}")

    def generate_reconciliation_report(
        self, df: pd.DataFrame, matches: List[Dict], output_path: str
    ) -> Tuple[bool, str]:
        """
        Generate comprehensive reconciliation report

        Args:
            df: DataFrame with reconciliation data
            matches: List of matched groups
            output_path: Path for report file

        Returns:
            Tuple of (success, message)
        """
        try:
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                # Summary sheet
                summary_data = self._create_summary_data(df, matches)
                summary_df = pd.DataFrame([summary_data])
                summary_df.to_excel(writer, sheet_name="Summary", index=False)

                # Matched transactions sheet
                matched_df = df[df["match_status"] == "matched"]
                if not matched_df.empty:
                    matched_df.to_excel(writer, sheet_name="Matched", index=False)

                # Unmatched transactions sheet
                unmatched_df = df[df["match_status"] == "unmatched"]
                if not unmatched_df.empty:
                    unmatched_df.to_excel(writer, sheet_name="Unmatched", index=False)

                # Match details sheet
                match_details = self._create_match_details(matches)
                match_details_df = pd.DataFrame(match_details)
                if not match_details_df.empty:
                    match_details_df.to_excel(
                        writer, sheet_name="Match Details", index=False
                    )

            return True, f"Reconciliation report created: {output_path}"

        except Exception as e:
            return False, f"Error generating report: {str(e)}"

    def _create_summary_data(self, df: pd.DataFrame, matches: List[Dict]) -> Dict:
        """
        Create summary statistics for report

        Args:
            df: DataFrame with reconciliation data
            matches: List of matched groups

        Returns:
            Dictionary with summary data
        """
        expenses = df[df["transaction_type"] == "expense"]
        revenues = df[df["transaction_type"] == "revenue"]
        matched = df[df["match_status"] == "matched"]
        unmatched = df[df["match_status"] == "unmatched"]

        return {
            "Report Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Total Transactions": len(df),
            "Total Expenses": len(expenses),
            "Total Revenues": len(revenues),
            "Matched Transactions": len(matched),
            "Unmatched Transactions": len(unmatched),
            "Match Rate (%)": round(
                (len(matched) / len(df) * 100) if len(df) > 0 else 0, 2
            ),
            "Total Match Groups": len(matches),
            "Total Expense Amount": abs(expenses["amount"].sum()),
            "Total Revenue Amount": revenues["amount"].sum(),
            "Net Balance": df["amount"].sum(),
        }

    def _create_match_details(self, matches: List[Dict]) -> List[Dict]:
        """
        Create detailed match information

        Args:
            matches: List of matched groups

        Returns:
            List of match detail dictionaries
        """
        details = []

        for idx, match in enumerate(matches):
            match_group_id = f"MG_{idx:04d}"
            revenue = match.get("revenue", {})
            expenses = match.get("expenses", [])
            confidence = match.get("confidence", 0.0)
            match_type = match.get("match_type", "unknown")

            revenue_amount = abs(revenue.get("amount", 0)) if revenue else 0
            expense_total = sum(abs(exp.get("amount", 0)) for exp in expenses)
            balance = revenue_amount - expense_total

            details.append(
                {
                    "Match Group ID": match_group_id,
                    "Match Type": match_type,
                    "Confidence (%)": round(confidence * 100, 1),
                    "Revenue Amount": revenue_amount,
                    "Expense Count": len(expenses),
                    "Total Expenses": expense_total,
                    "Balance": balance,
                    "Status": "Balanced" if abs(balance) < 0.01 else "Unbalanced",
                }
            )

        return details

    def export_unmatched_only(
        self, df: pd.DataFrame, output_path: str
    ) -> Tuple[bool, str]:
        """
        Export only unmatched transactions for review

        Args:
            df: DataFrame with reconciliation data
            output_path: Path for output file

        Returns:
            Tuple of (success, message)
        """
        try:
            unmatched = df[df["match_status"] == "unmatched"]

            if unmatched.empty:
                return False, "No unmatched transactions to export"

            # Separate expenses and revenues
            unmatched_expenses = unmatched[unmatched["transaction_type"] == "expense"]
            unmatched_revenues = unmatched[unmatched["transaction_type"] == "revenue"]

            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                unmatched_expenses.to_excel(
                    writer, sheet_name="Unmatched Expenses", index=False
                )
                unmatched_revenues.to_excel(
                    writer, sheet_name="Unmatched Revenues", index=False
                )

            return True, f"Unmatched transactions exported: {output_path}"

        except Exception as e:
            return False, f"Error exporting unmatched transactions: {str(e)}"
